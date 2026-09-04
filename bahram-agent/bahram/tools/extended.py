"""
Extended.

Public objects: ``GitTool``, ``ProcessListTool``, ``ContainerTool``, ``DocumentReadTool``.
"""

from __future__ import annotations

import asyncio
import logging
from typing import Any

from bahram.tools.base import BaseTool

logger = logging.getLogger(__name__)


class GitTool(BaseTool):
    """
    Git tool.
    """

    @property
    def name(self) -> str:
        """
        Return the registry name of the GitTool object.

        Returns the constant string ``'git'``.

        Returns:
            str: the rendered string.
        """
        return "git"

    @property
    def description(self) -> str:
        """
        Return the human readable description shown to the model.

        Returns the constant string ``'Execute git commands (status, log, diff, add, commit, branch,
            checkout, etc.).'``.

        Returns:
            str: the rendered string.
        """
        return "Execute git commands (status, log, diff, add, commit, branch, checkout, etc.)."

    @property
    def parameters(self) -> dict[str, Any]:
        """
        Return the JSON schema describing this tool's arguments.

        Returns:
            dict[str, Any]: a mapping of str, Any.
        """
        return {
            "type": "object",
            "properties": {
                "command": {
                    "type": "string",
                    "description": (
                        "Git command to run (e.g. 'status', 'log --oneline -5', 'diff', 'add .')"
                    ),
                },
                "workdir": {
                    "type": "string",
                    "description": "Working directory (must be inside a git repo)",
                },
            },
            "required": ["command"],
        }

    async def execute(self, **kwargs: Any) -> str:
        """
        Execute the tool and return its textual result.

        Args:
            **kwargs (Any): keyword arguments forwarded to the implementation.

        Returns:
            str: the rendered string.

        Note:
            Coroutine - must be awaited.
        """
        command = kwargs.get("command", "")
        workdir = kwargs.get("workdir", None)

        if not command:
            return "Error: No git command provided"

        full_cmd = f"git {command}"

        try:
            process = await asyncio.create_subprocess_shell(
                full_cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
                cwd=workdir,
            )
            stdout, stderr = await asyncio.wait_for(process.communicate(), timeout=60)
            stdout_str = stdout.decode("utf-8", errors="replace")
            stderr_str = stderr.decode("utf-8", errors="replace")

            if process.returncode != 0:
                return f"Git error: {stderr_str or 'unknown error'}"
            return stdout_str or stderr_str or "(no output)"
        except asyncio.TimeoutError:
            return "Error: Git command timed out"
        except Exception as e:
            return f"Error: {e}"


class ProcessListTool(BaseTool):
    """
    Process list tool.
    """

    @property
    def name(self) -> str:
        """
        Return the registry name of the ProcessListTool object.

        Returns the constant string ``'process_list'``.

        Returns:
            str: the rendered string.
        """
        return "process_list"

    @property
    def description(self) -> str:
        """
        Return the human readable description shown to the model.

        Returns the constant string ``'List running processes or get info about a specific
            process.'``.

        Returns:
            str: the rendered string.
        """
        return "List running processes or get info about a specific process."

    @property
    def parameters(self) -> dict[str, Any]:
        """
        Return the JSON schema describing this tool's arguments.

        Returns:
            dict[str, Any]: a mapping of str, Any.
        """
        return {
            "type": "object",
            "properties": {
                "pid": {
                    "type": "integer",
                    "description": "PID to inspect (optional, lists all if omitted)",
                },
                "filter": {
                    "type": "string",
                    "description": "Filter by process name pattern",
                },
            },
        }

    async def execute(self, **kwargs: Any) -> str:
        """
        Execute the tool and return its textual result.

        Args:
            **kwargs (Any): keyword arguments forwarded to the implementation.

        Returns:
            str: the rendered string.

        Note:
            Coroutine - must be awaited.
        """
        pid = kwargs.get("pid")
        proc_filter = kwargs.get("filter", "")

        if pid:
            try:
                with open(f"/proc/{pid}/status") as f:
                    return f.read()[:2000]
            except FileNotFoundError:
                return f"Error: Process {pid} not found"
            except Exception as e:
                return f"Error: {e}"

        cmd = "ps aux"
        if proc_filter:
            cmd = f"ps aux | grep {proc_filter} | grep -v grep"

        process = await asyncio.create_subprocess_shell(
            cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        stdout, _ = await process.communicate()
        return stdout.decode("utf-8", errors="replace")[:3000]


class ContainerTool(BaseTool):
    """
    Container tool.
    """

    @property
    def name(self) -> str:
        """
        Return the registry name of the ContainerTool object.

        Returns the constant string ``'container'``.

        Returns:
            str: the rendered string.
        """
        return "container"

    @property
    def description(self) -> str:
        """
        Return the human readable description shown to the model.

        Returns the constant string ``'Run container operations (list, inspect, exec, logs,
            stats).'``.

        Returns:
            str: the rendered string.
        """
        return "Run container operations (list, inspect, exec, logs, stats)."

    @property
    def parameters(self) -> dict[str, Any]:
        """
        Return the JSON schema describing this tool's arguments.

        Returns:
            dict[str, Any]: a mapping of str, Any.
        """
        return {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["list", "inspect", "exec", "logs", "stats"],
                    "description": "Container action",
                },
                "container": {
                    "type": "string",
                    "description": "Container name or ID",
                },
                "command": {
                    "type": "string",
                    "description": "Command to exec (for exec action)",
                },
            },
            "required": ["action"],
        }

    async def execute(self, **kwargs: Any) -> str:
        """
        Execute the tool and return its textual result.

        Args:
            **kwargs (Any): keyword arguments forwarded to the implementation.

        Returns:
            str: the rendered string.

        Note:
            Coroutine - must be awaited.
        """
        action = kwargs.get("action", "")
        container = kwargs.get("container", "")
        command = kwargs.get("command", "")

        if action == "list":
            cmd = "docker ps -a --format 'table {{.ID}}\t{{.Names}}\t{{.Status}}\t{{.Image}}'"
        elif action == "inspect" and container:
            cmd = f"docker inspect {container}"
        elif action == "logs" and container:
            cmd = f"docker logs --tail 100 {container}"
        elif action == "stats":
            cmd = "docker stats --no-stream --format 'table {{.Name}}\t{{.CPUPerc}}\t{{.MemUsage}}'"
        elif action == "exec" and container and command:
            cmd = f"docker exec {container} {command}"
        else:
            return "Error: Invalid action or missing container/command"

        try:
            process = await asyncio.create_subprocess_shell(
                cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await asyncio.wait_for(process.communicate(), timeout=30)
            if process.returncode != 0:
                return f"Docker error: {stderr.decode()}"
            return stdout.decode() or "(no output)"
        except asyncio.TimeoutError:
            return "Error: Docker command timed out"
        except Exception as e:
            return f"Error: {e}"


class DocumentReadTool(BaseTool):
    """
    Document read tool.
    """

    @property
    def name(self) -> str:
        """
        Return the registry name of the DocumentReadTool object.

        Returns the constant string ``'document_read'``.

        Returns:
            str: the rendered string.
        """
        return "document_read"

    @property
    def description(self) -> str:
        """
        Return the human readable description shown to the model.

        Returns the constant string ``'Read and extract text from documents (PDF, DOCX, TXT, MD,
            CSV, JSON, YAML).'``.

        Returns:
            str: the rendered string.
        """
        return "Read and extract text from documents (PDF, DOCX, TXT, MD, CSV, JSON, YAML)."

    @property
    def parameters(self) -> dict[str, Any]:
        """
        Return the JSON schema describing this tool's arguments.

        Returns:
            dict[str, Any]: a mapping of str, Any.
        """
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the document",
                },
                "format": {
                    "type": "string",
                    "description": "Force format (auto-detected if omitted)",
                },
            },
            "required": ["file_path"],
        }

    async def execute(self, **kwargs: Any) -> str:
        """
        Execute the tool and return its textual result.

        Args:
            **kwargs (Any): keyword arguments forwarded to the implementation.

        Returns:
            str: the rendered string.

        Note:
            Coroutine - must be awaited.
        """
        import os

        file_path = kwargs.get("file_path", "")

        if not file_path:
            return "Error: No file path provided"

        if not os.path.exists(file_path):
            return f"Error: File not found: {file_path}"

        ext = os.path.splitext(file_path)[1].lower()
        max_size = 10 * 1024 * 1024

        if os.path.getsize(file_path) > max_size:
            return "Error: File too large (>10MB)"

        try:
            if ext in (
                ".txt",
                ".md",
                ".py",
                ".js",
                ".ts",
                ".json",
                ".yaml",
                ".yml",
                ".toml",
                ".cfg",
                ".ini",
                ".csv",
                ".log",
            ):
                with open(file_path, encoding="utf-8", errors="replace") as f:
                    content = f.read(50000)
                return content

            elif ext == ".pdf":
                try:
                    import subprocess

                    result = subprocess.run(
                        ["pdftotext", file_path, "-"],
                        capture_output=True,
                        text=True,
                        timeout=30,
                    )
                    return (
                        result.stdout[:50000]
                        if result.returncode == 0
                        else f"Error: {result.stderr}"
                    )
                except FileNotFoundError:
                    return "Error: pdftotext not installed"
                except Exception as e:
                    return f"Error reading PDF: {e}"

            elif ext == ".docx":
                try:
                    from docx import Document

                    doc = Document(file_path)
                    return "\n".join(p.text for p in doc.paragraphs)[:50000]
                except ImportError:
                    return "Error: python-docx not installed"
                except Exception as e:
                    return f"Error reading DOCX: {e}"

            elif ext == ".xlsx":
                try:
                    from openpyxl import load_workbook

                    wb = load_workbook(file_path, read_only=True)
                    lines = []
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        lines.append(f"=== Sheet: {sheet} ===")
                        for row in ws.iter_rows(max_row=100, values_only=True):
                            lines.append("\t".join(str(c or "") for c in row))
                    wb.close()
                    return "\n".join(lines)[:50000]
                except ImportError:
                    return "Error: openpyxl not installed"
                except Exception as e:
                    return f"Error reading XLSX: {e}"

            else:
                try:
                    with open(file_path, encoding="utf-8", errors="replace") as f:
                        return f.read(50000)
                except Exception as e:
                    return f"Error reading file: {e}"

        except Exception as e:
            return f"Error: {e}"
